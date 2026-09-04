"""Integration — opening batch S1: GL+bank → lock → voucher gate."""

from __future__ import annotations

from datetime import date
from uuid import UUID, uuid4

import pytest

from tests.integration.conftest import (
    UUID_AUDITOR,
    UUID_CHIEF,
    FakeUser,
    _store,
)

COMPANY = "99999999-9999-9999-9999-999999999999"


@pytest.fixture()
def chief(app):
    u = FakeUser(UUID_CHIEF, "CHIEF_ACCOUNTANT")
    _store[u.id] = u
    c = app.test_client()
    with c.session_transaction() as sess:
        sess["_user_id"] = u.id
    return c


@pytest.fixture()
def auditor(app):
    u = FakeUser(UUID_AUDITOR, "AUDITOR")
    _store[u.id] = u
    c = app.test_client()
    with c.session_transaction() as sess:
        sess["_user_id"] = u.id
    return c


@pytest.fixture()
def ready(app):
    fy, _periods = app.fy_service.create_year(
        UUID(COMPANY),
        "2026",
        date(2026, 1, 1),
        date(2026, 12, 31),
        "MONTHLY",
        actor=uuid4(),
        reason="fy",
    )
    coa = app.coa_service
    coa.create_account(UUID(COMPANY), "111", "Tiền mặt", actor=uuid4(), reason="c")
    coa.create_account(
        UUID(COMPANY), "1111", "Tiền mặt VNĐ", parent_code="111", actor=uuid4(), reason="c"
    )
    coa.create_account(UUID(COMPANY), "411", "Vốn", actor=uuid4(), reason="c")
    coa.create_account(
        UUID(COMPANY), "4111", "Vốn góp", parent_code="411", actor=uuid4(), reason="c"
    )
    coa.create_account(UUID(COMPANY), "131", "PTKH", actor=uuid4(), reason="c")
    coa.create_account(
        UUID(COMPANY), "1311", "PTKH ct", parent_code="131", actor=uuid4(), reason="c"
    )
    coa.create_account(UUID(COMPANY), "152", "NVL", actor=uuid4(), reason="c")
    coa.create_account(
        UUID(COMPANY), "1521", "NVL ct", parent_code="152", actor=uuid4(), reason="c"
    )
    coa.create_account(UUID(COMPANY), "112", "TGNH", actor=uuid4(), reason="c")
    coa.create_account(
        UUID(COMPANY), "1121", "TGNH VNĐ", parent_code="112", actor=uuid4(), reason="c"
    )
    coa.create_account(UUID(COMPANY), "211", "TSCĐ", actor=uuid4(), reason="c")
    coa.create_account(
        UUID(COMPANY), "2111", "TSCĐ HH", parent_code="211", actor=uuid4(), reason="c"
    )
    coa.create_account(UUID(COMPANY), "214", "Hao mòn", actor=uuid4(), reason="c")
    coa.create_account(
        UUID(COMPANY), "2141", "HM TSCĐ HH", parent_code="214", actor=uuid4(), reason="c"
    )
    coa.create_account(UUID(COMPANY), "642", "QLDN", actor=uuid4(), reason="c")
    coa.create_account(
        UUID(COMPANY), "6421", "QLDN ct", parent_code="642", actor=uuid4(), reason="c"
    )
    coa.create_account(UUID(COMPANY), "242", "CPTT", actor=uuid4(), reason="c")
    coa.create_account(
        UUID(COMPANY), "2421", "CPTT ct", parent_code="242", actor=uuid4(), reason="c"
    )
    coa.create_account(UUID(COMPANY), "627", "SXC", actor=uuid4(), reason="c")
    from src.bricks.payment_terms.web_adapter import _series_service as ss

    ss.create_series(company_id=UUID(COMPANY), prefix="PT/", actor=uuid4(), reason="s")
    return {"fy_id": str(fy.id)}


VOUCHER = {
    "company_id": COMPANY,
    "entry_date": "2026-08-12",
    "description": "Thu tiền",
    "lines": [
        {"account_code": "1111", "debit": "100", "credit": "0"},
        {"account_code": "4111", "debit": "0", "credit": "100"},
    ],
}


class TestOpeningFlow:
    def test_full_s1_flow(self, chief, ready):
        fy_id = ready["fy_id"]
        r = chief.post(
            "/api/v1/opening-batches",
            json={"company_id": COMPANY, "fiscal_year_id": fy_id, "reason": "init"},
        )
        assert r.status_code == 201, r.get_json()
        bid = r.get_json()["data"]["id"]

        # live voucher blocked before lock
        blocked = chief.post("/api/v1/vouchers", json=VOUCHER)
        assert blocked.status_code == 409
        assert blocked.get_json()["code"] == "NO_OPENING_LOCK"

        # lock unbalanced → 409
        chief.post(
            f"/api/v1/opening-batches/{bid}/gl",
            json={
                "reason": "gl",
                "lines": [
                    {"account_code": "1111", "debit": "500", "credit": "0"},
                    {"account_code": "4111", "debit": "0", "credit": "400"},
                ],
            },
        )
        assert (
            chief.post(f"/api/v1/opening-batches/{bid}/lock", json={"reason": "go"}).status_code
            == 409
        )

        # fix first batch via balancing GL post, then lock it too:
        # hardened gate requires EVERY batch locked, not just one
        chief.post(
            f"/api/v1/opening-batches/{bid}/gl",
            json={
                "reason": "gl-fix",
                "lines": [{"account_code": "4111", "debit": "0", "credit": "100"}],
            },
        )
        fix1 = chief.post(f"/api/v1/opening-batches/{bid}/lock", json={"reason": "go"})
        assert fix1.status_code == 200

        # second batch carries bank tie + full flow
        r2 = chief.post(
            "/api/v1/opening-batches",
            json={"company_id": COMPANY, "fiscal_year_id": fy_id, "reason": "init2"},
        )
        bid2 = r2.get_json()["data"]["id"]
        chief.post(
            f"/api/v1/opening-batches/{bid2}/gl",
            json={
                "reason": "gl",
                "lines": [
                    {"account_code": "1121", "debit": "500", "credit": "0"},
                    {"account_code": "4111", "debit": "0", "credit": "500"},
                ],
            },
        )
        chief.post(
            f"/api/v1/opening-batches/{bid2}/bank",
            json={
                "reason": "bank",
                "rows": [{"bank_account_id": str(uuid4()), "amount": "500"}],
            },
        )
        rep = chief.get(f"/api/v1/opening-batches/{bid2}/reconcile").get_json()["data"]
        assert rep["balanced"] is True
        assert rep["checks"]["bank_total"] == 500.0

        lock = chief.post(f"/api/v1/opening-batches/{bid2}/lock", json={"reason": "go-live"})
        assert lock.status_code == 200
        assert lock.get_json()["data"]["state"] == "LOCKED"

        # live voucher now allowed
        ok = chief.post("/api/v1/vouchers", json=VOUCHER)
        assert ok.status_code == 201, ok.get_json()

    def test_auditor_read_only(self, chief, auditor, ready):
        fy_id = ready["fy_id"]
        r = chief.post(
            "/api/v1/opening-batches",
            json={"company_id": COMPANY, "fiscal_year_id": fy_id, "reason": "init"},
        )
        assert r.status_code == 201
        denied = auditor.post(
            "/api/v1/opening-batches",
            json={"company_id": COMPANY, "fiscal_year_id": fy_id, "reason": "x"},
        )
        assert denied.status_code == 403
        assert (
            auditor.get(
                f"/api/v1/opening-batches/{r.get_json()['data']['id']}/reconcile"
            ).status_code
            == 200
        )


class TestCounterpartyFlow:
    def test_counterparty_tie_and_aging(self, chief, ready):
        fy_id = ready["fy_id"]
        p = chief.post(
            "/api/v1/parties",
            json={
                "company_id": COMPANY,
                "code": "KH-001",
                "name": "Khách A",
                "mst": "0101234567",
                "is_customer": True,
            },
        )
        assert p.status_code == 201, p.get_json()
        pid = p.get_json()["data"]["id"]

        r = chief.post(
            "/api/v1/opening-batches",
            json={"company_id": COMPANY, "fiscal_year_id": fy_id, "reason": "init"},
        )
        bid = r.get_json()["data"]["id"]
        cp = chief.post(
            f"/api/v1/opening-batches/{bid}/counterparties",
            json={
                "reason": "ar",
                "rows": [
                    {
                        "account_code": "1311",
                        "party_id": pid,
                        "side": "debit",
                        "amount": "200",
                    }
                ],
            },
        )
        assert cp.status_code == 201, cp.get_json()

        # GL must tie or lock fails
        chief.post(
            f"/api/v1/opening-batches/{bid}/gl",
            json={
                "reason": "gl",
                "lines": [
                    {"account_code": "1311", "debit": "200", "credit": "0"},
                    {"account_code": "4111", "debit": "0", "credit": "200"},
                ],
            },
        )
        # need 4111 aggregate? use detail accounts instead
        lock = chief.post(f"/api/v1/opening-batches/{bid}/lock", json={"reason": "go"})
        assert lock.status_code == 200, lock.get_json()

        aging = chief.get(
            "/api/v1/reports/ar-aging",
            query_string={"company_id": COMPANY, "as_of": "2026-08-31"},
        )
        buckets = {x["bucket"]: x["amount"] for x in aging.get_json()["data"]}
        assert buckets["current"] == 200.0


class TestStockFlow:
    def test_stock_tie_and_nxt(self, chief, ready):
        fy_id = ready["fy_id"]
        cat = chief.post(
            "/api/v1/inventory/categories",
            json={
                "company_id": COMPANY,
                "code": "CAT-OP",
                "name": "NVL OP",
                "account_code": "1521",
            },
        ).get_json()["data"]["id"]
        prod = chief.post(
            "/api/v1/inventory/products",
            json={
                "company_id": COMPANY,
                "code": "SKU-OP",
                "name": "NVL OP",
                "uom": "Cái",
                "cost_method": "wavg",
                "category_id": cat,
            },
        ).get_json()["data"]["id"]
        wh = chief.post(
            "/api/v1/inventory/warehouses",
            json={"company_id": COMPANY, "code": "KHO-OP", "name": "Kho OP"},
        ).get_json()["data"]["id"]
        loc = chief.post(
            "/api/v1/inventory/locations",
            json={"company_id": COMPANY, "code": "A-OP", "name": "Kệ OP", "warehouse_id": wh},
        ).get_json()["data"]["id"]

        r = chief.post(
            "/api/v1/opening-batches",
            json={"company_id": COMPANY, "fiscal_year_id": fy_id, "reason": "init"},
        )
        bid = r.get_json()["data"]["id"]
        st = chief.post(
            f"/api/v1/opening-batches/{bid}/stock",
            json={
                "reason": "stock",
                "rows": [
                    {
                        "product_id": prod,
                        "warehouse_id": loc,
                        "qty": "100",
                        "total_value": "1000000",
                    }
                ],
            },
        )
        assert st.status_code == 201, st.get_json()

        # tie mismatch blocks lock
        chief.post(
            f"/api/v1/opening-batches/{bid}/gl",
            json={
                "reason": "gl",
                "lines": [
                    {"account_code": "1521", "debit": "900000", "credit": "0"},
                    {"account_code": "4111", "debit": "0", "credit": "900000"},
                ],
            },
        )
        bad = chief.post(f"/api/v1/opening-batches/{bid}/lock", json={"reason": "go"})
        assert bad.status_code == 409
        assert "1521" in bad.get_json()["error"]

        # fix GL via second batch is overkill — reopen path covered in unit;
        # verify NXT sees the materialized move instead
        nxt = chief.get(
            "/api/v1/reports/inventory/nxt",
            query_string={"company_id": COMPANY, "from": "2026-01-01", "to": "2026-12-31"},
        )
        rows = {x["code"]: x for x in nxt.get_json()["data"]}
        assert rows["SKU-OP"]["in_qty"] == 100.0


class TestAssetFlow:
    def test_asset_posts_and_materializes_fa(self, chief, ready):
        fy_id = ready["fy_id"]
        r = chief.post(
            "/api/v1/opening-batches",
            json={"company_id": COMPANY, "fiscal_year_id": fy_id, "reason": "init"},
        )
        bid = r.get_json()["data"]["id"]
        st = chief.post(
            f"/api/v1/opening-batches/{bid}/assets",
            json={
                "reason": "fa",
                "rows": [
                    {
                        "kind": "fixed_asset",
                        "code": "TSCD-OP",
                        "name": "Máy CNC",
                        "original_cost": "1200000000",
                        "remaining_value": "800000000",
                        "months_left": 80,
                        "expense_account": "6421",
                    }
                ],
            },
        )
        assert st.status_code == 201, st.get_json()

        fa = chief.get("/api/v1/fixed-assets", query_string={"company_id": COMPANY}).get_json()[
            "data"
        ]
        mine = next(x for x in fa if x["asset_code"] == "TSCD-OP")
        assert mine["accumulated_depreciation"] == 400000000.0
        assert mine["book_value"] == 800000000.0

        # FA tie mismatch blocks lock (211 net 1.2B vs ... here GL missing 214)
        chief.post(
            f"/api/v1/opening-batches/{bid}/gl",
            json={
                "reason": "gl",
                "lines": [
                    {"account_code": "2111", "debit": "1200000000", "credit": "0"},
                    {"account_code": "4111", "debit": "0", "credit": "1200000000"},
                ],
            },
        )
        bad = chief.post(f"/api/v1/opening-batches/{bid}/lock", json={"reason": "go"})
        assert bad.status_code == 409
        assert "211" in bad.get_json()["error"]


class TestExcelImport:
    def _workbook(self, rows):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["account_code", "debit", "credit"])
        for r in rows:
            ws.append(r)
        import io

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def test_gl_excel_import_posts_and_locks(self, chief, ready):
        fy_id = ready["fy_id"]
        r = chief.post(
            "/api/v1/opening-batches",
            json={"company_id": COMPANY, "fiscal_year_id": fy_id, "reason": "init"},
        )
        bid = r.get_json()["data"]["id"]
        buf = self._workbook([["1111", 500, 0], ["4111", 0, 500]])
        imp = chief.post(
            f"/api/v1/opening-batches/{bid}/gl/import",
            data={"reason": "excel", "file": (buf, "gl.xlsx")},
            content_type="multipart/form-data",
        )
        assert imp.status_code == 201, imp.get_json()
        assert imp.get_json()["data"]["lines"] == 2
        lock = chief.post(f"/api/v1/opening-batches/{bid}/lock", json={"reason": "go"})
        assert lock.status_code == 200, lock.get_json()

    def test_gl_excel_import_rejects_bad_header(self, chief, ready):
        fy_id = ready["fy_id"]
        r = chief.post(
            "/api/v1/opening-batches",
            json={"company_id": COMPANY, "fiscal_year_id": fy_id, "reason": "init"},
        )
        bid = r.get_json()["data"]["id"]
        buf = self._workbook([["1111", 500, 0]])
        # corrupt header
        from openpyxl import load_workbook

        wb = load_workbook(buf)
        wb.active["A1"] = "tk"
        import io

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        imp = chief.post(
            f"/api/v1/opening-batches/{bid}/gl/import",
            data={"reason": "excel", "file": (out, "gl.xlsx")},
            content_type="multipart/form-data",
        )
        assert imp.status_code == 422


class TestCCDCFlow:
    def test_ccdc_posts_backfills_and_ties_242(self, chief, ready):
        fy_id = ready["fy_id"]
        r = chief.post(
            "/api/v1/opening-batches",
            json={"company_id": COMPANY, "fiscal_year_id": fy_id, "reason": "init"},
        )
        bid = r.get_json()["data"]["id"]
        st = chief.post(
            f"/api/v1/opening-batches/{bid}/assets",
            json={
                "reason": "ccdc",
                "rows": [
                    {
                        "kind": "ccdc",
                        "code": "CCDC-OP",
                        "name": "Máy khoan",
                        "original_cost": "12000000",
                        "remaining_value": "7000000",
                        "months_left": 7,
                        "useful_life_months": 12,
                        "expense_account": "627",
                        "purchase_date": "2026-01-15",
                    }
                ],
            },
        )
        assert st.status_code == 201, st.get_json()

        # 242 mismatch blocks lock
        chief.post(
            f"/api/v1/opening-batches/{bid}/gl",
            json={
                "reason": "gl",
                "lines": [
                    {"account_code": "2421", "debit": "5000000", "credit": "0"},
                    {"account_code": "4111", "debit": "0", "credit": "5000000"},
                ],
            },
        )
        bad = chief.post(f"/api/v1/opening-batches/{bid}/lock", json={"reason": "go"})
        assert bad.status_code == 409
        assert "242" in bad.get_json()["error"]
