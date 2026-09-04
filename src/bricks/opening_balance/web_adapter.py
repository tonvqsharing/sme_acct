"""Opening balance web adapter — ONLY Flask file in brick."""

from __future__ import annotations

from typing import Any
from uuid import UUID

from flask import Blueprint, abort, jsonify, request
from flask_login import current_user, login_required

from src.bricks.opening_balance.services import (
    BatchLockedError,
    NotFoundError,
    UnbalancedOpeningError,
)

opening_balance_bp = Blueprint("opening_balance", __name__)

_opening_service: Any = None


def init_opening_service(svc: Any) -> None:
    global _opening_service
    _opening_service = svc


def _svc() -> Any:
    s = _opening_service
    if s is None:
        abort(500, description="OpeningService not initialized")
    return s


def _require_write() -> None:
    role = getattr(current_user, "role", "")
    if role == "AUDITOR":
        abort(403, description="AUDITOR chỉ đọc")
    if role not in ("ADMIN", "ACCOUNTANT", "CHIEF_ACCOUNTANT", "DIRECTOR"):
        abort(403)


def _require_chief() -> None:
    role = getattr(current_user, "role", "")
    if role not in ("CHIEF_ACCOUNTANT", "ADMIN"):
        abort(403, description="Only CHIEF_ACCOUNTANT/ADMIN")


@opening_balance_bp.post("/api/v1/opening-batches")
@login_required  # type: ignore[untyped-decorator]
def create_batch() -> tuple[Any, int]:
    _require_write()
    body = request.get_json(silent=True) or {}
    try:
        b = _svc().create_batch(
            company_id=UUID(body["company_id"]),
            fiscal_year_id=UUID(body["fiscal_year_id"]),
            source=body.get("source", "MANUAL"),
            actor=UUID(str(current_user.id)),
            reason=body.get("reason") or "create opening batch",
        )
    except (KeyError, ValueError) as exc:
        return jsonify({"error": str(exc), "code": "INVALID_BATCH"}), 422
    except NotFoundError as exc:
        return jsonify({"error": str(exc), "code": "NOT_FOUND"}), 404
    return jsonify({"data": {"id": str(b.id), "state": b.state.value}}), 201


@opening_balance_bp.post("/api/v1/opening-batches/<bid>/gl")
@login_required  # type: ignore[untyped-decorator]
def post_gl(bid: str) -> tuple[Any, int]:
    _require_write()
    body = request.get_json(silent=True) or {}
    try:
        bid_u = UUID(bid)
    except ValueError:
        abort(422, description="Invalid UUID")
    try:
        _svc().post_gl(
            bid_u,
            lines=body.get("lines", []),
            actor=UUID(str(current_user.id)),
            reason=body.get("reason") or "post gl opening",
        )
    except (KeyError, ValueError) as exc:
        return jsonify({"error": str(exc), "code": "INVALID_GL"}), 422
    except BatchLockedError as exc:
        return jsonify({"error": str(exc), "code": "BATCH_LOCKED"}), 409
    except NotFoundError as exc:
        return jsonify({"error": str(exc), "code": "NOT_FOUND"}), 404
    return jsonify({"data": {"posted": True}}), 201


@opening_balance_bp.post("/api/v1/opening-batches/<bid>/gl/import")
@login_required  # type: ignore[untyped-decorator]
def import_gl_excel(bid: str) -> tuple[Any, int]:
    """Upload .xlsx (header account_code/debit/credit[, currency_code]) → post_gl."""
    _require_write()
    try:
        bid_u = UUID(bid)
    except ValueError:
        abort(422, description="Invalid UUID")
    upload = request.files.get("file")
    if upload is None or not (upload.filename or "").endswith((".xlsx", ".xlsm")):
        return jsonify({"error": "file .xlsx required", "code": "INVALID_FILE"}), 422
    try:
        from openpyxl import load_workbook

        wb = load_workbook(upload, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return jsonify({"error": "empty workbook", "code": "EMPTY_SHEET"}), 422
        header = [str(c.value or "").strip().lower() for c in next(ws.rows)]
        required = ["account_code", "debit", "credit"]
        if header[:3] != required:
            return jsonify({"error": f"header must be {required}", "code": "BAD_HEADER"}), 422
        lines = []
        for row in list(ws.rows)[1:]:
            code = str(row[0].value or "").strip()
            if not code:
                continue
            lines.append(
                {
                    "account_code": code,
                    "debit": str(row[1].value or 0),
                    "credit": str(row[2].value or 0),
                    **(
                        {"currency_code": str(row[3].value).strip()}
                        if len(row) > 3 and row[3].value
                        else {}
                    ),
                }
            )
    except (KeyError, ValueError) as exc:
        return jsonify({"error": str(exc), "code": "INVALID_GL"}), 422
    if not lines:
        return jsonify({"error": "no data rows", "code": "EMPTY_SHEET"}), 422
    try:
        _svc().post_gl(
            bid_u,
            lines=lines,
            actor=UUID(str(current_user.id)),
            reason=request.form.get("reason") or "excel import gl",
        )
    except (KeyError, ValueError) as exc:
        return jsonify({"error": str(exc), "code": "INVALID_GL"}), 422
    except BatchLockedError as exc:
        return jsonify({"error": str(exc), "code": "BATCH_LOCKED"}), 409
    except NotFoundError as exc:
        return jsonify({"error": str(exc), "code": "NOT_FOUND"}), 404
    return jsonify({"data": {"posted": True, "lines": len(lines)}}), 201


@opening_balance_bp.post("/api/v1/opening-batches/<bid>/bank")
@login_required  # type: ignore[untyped-decorator]
def post_bank(bid: str) -> tuple[Any, int]:
    _require_write()
    body = request.get_json(silent=True) or {}
    try:
        bid_u = UUID(bid)
    except ValueError:
        abort(422, description="Invalid UUID")
    try:
        _svc().post_bank(
            bid_u,
            rows=body.get("rows", []),
            actor=UUID(str(current_user.id)),
            reason=body.get("reason") or "post bank opening",
        )
    except (KeyError, ValueError) as exc:
        return jsonify({"error": str(exc), "code": "INVALID_BANK"}), 422
    except BatchLockedError as exc:
        return jsonify({"error": str(exc), "code": "BATCH_LOCKED"}), 409
    except NotFoundError as exc:
        return jsonify({"error": str(exc), "code": "NOT_FOUND"}), 404
    return jsonify({"data": {"posted": True}}), 201


@opening_balance_bp.post("/api/v1/opening-batches/<bid>/counterparties")
@login_required  # type: ignore[untyped-decorator]
def post_counterparty(bid: str) -> tuple[Any, int]:
    _require_write()
    body = request.get_json(silent=True) or {}
    try:
        bid_u = UUID(bid)
    except ValueError:
        abort(422, description="Invalid UUID")
    try:
        _svc().post_counterparty(
            bid_u,
            rows=body.get("rows", []),
            actor=UUID(str(current_user.id)),
            reason=body.get("reason") or "post counterparty opening",
        )
    except (KeyError, ValueError) as exc:
        return jsonify({"error": str(exc), "code": "INVALID_COUNTERPARTY"}), 422
    except BatchLockedError as exc:
        return jsonify({"error": str(exc), "code": "BATCH_LOCKED"}), 409
    except NotFoundError as exc:
        return jsonify({"error": str(exc), "code": "NOT_FOUND"}), 404
    return jsonify({"data": {"posted": True}}), 201


@opening_balance_bp.post("/api/v1/opening-batches/<bid>/stock")
@login_required  # type: ignore[untyped-decorator]
def post_stock(bid: str) -> tuple[Any, int]:
    _require_write()
    body = request.get_json(silent=True) or {}
    try:
        bid_u = UUID(bid)
    except ValueError:
        abort(422, description="Invalid UUID")
    try:
        _svc().post_stock(
            bid_u,
            rows=body.get("rows", []),
            actor=UUID(str(current_user.id)),
            reason=body.get("reason") or "post stock opening",
        )
    except (KeyError, ValueError) as exc:
        return jsonify({"error": str(exc), "code": "INVALID_STOCK"}), 422
    except BatchLockedError as exc:
        return jsonify({"error": str(exc), "code": "BATCH_LOCKED"}), 409
    except NotFoundError as exc:
        return jsonify({"error": str(exc), "code": "NOT_FOUND"}), 404
    except RuntimeError as exc:
        return jsonify({"error": str(exc), "code": "NOT_WIRED"}), 500
    return jsonify({"data": {"posted": True}}), 201


@opening_balance_bp.post("/api/v1/opening-batches/<bid>/assets")
@login_required  # type: ignore[untyped-decorator]
def post_assets(bid: str) -> tuple[Any, int]:
    _require_write()
    body = request.get_json(silent=True) or {}
    try:
        bid_u = UUID(bid)
    except ValueError:
        abort(422, description="Invalid UUID")
    try:
        _svc().post_assets(
            bid_u,
            rows=body.get("rows", []),
            actor=UUID(str(current_user.id)),
            reason=body.get("reason") or "post asset opening",
        )
    except (KeyError, ValueError) as exc:
        return jsonify({"error": str(exc), "code": "INVALID_ASSET"}), 422
    except BatchLockedError as exc:
        return jsonify({"error": str(exc), "code": "BATCH_LOCKED"}), 409
    except NotFoundError as exc:
        return jsonify({"error": str(exc), "code": "NOT_FOUND"}), 404
    except RuntimeError as exc:
        return jsonify({"error": str(exc), "code": "NOT_WIRED"}), 500
    return jsonify({"data": {"posted": True}}), 201


@opening_balance_bp.get("/api/v1/opening-batches/<bid>/reconcile")
@login_required  # type: ignore[untyped-decorator]
def reconcile(bid: str) -> tuple[Any, int]:
    try:
        bid_u = UUID(bid)
    except ValueError:
        abort(422, description="Invalid UUID")
    try:
        rep = _svc().reconcile(bid_u)
    except NotFoundError as exc:
        return jsonify({"error": str(exc), "code": "NOT_FOUND"}), 404
    return (
        jsonify(
            {
                "data": {
                    "balanced": rep["balanced"],
                    "debit_total": float(rep["debit_total"]),
                    "credit_total": float(rep["credit_total"]),
                    "checks": {
                        "bank_total": float(rep["checks"]["bank_total"]),
                        "gl_lines": rep["checks"]["gl_lines"],
                        "counterparty_total": float(rep["checks"]["counterparty_total"]),
                        "counterparty_lines": rep["checks"]["counterparty_lines"],
                        "stock_total": float(rep["checks"]["stock_total"]),
                        "stock_lines": rep["checks"]["stock_lines"],
                    },
                }
            }
        ),
        200,
    )


@opening_balance_bp.post("/api/v1/opening-batches/<bid>/lock")
@login_required  # type: ignore[untyped-decorator]
def lock(bid: str) -> tuple[Any, int]:
    _require_chief()
    body = request.get_json(silent=True) or {}
    try:
        b = _svc().lock(
            UUID(bid),
            actor=UUID(str(current_user.id)),
            reason=body.get("reason") or "go-live",
        )
    except UnbalancedOpeningError as exc:
        return jsonify({"error": str(exc), "code": "UNBALANCED_OPENING"}), 409
    except BatchLockedError as exc:
        return jsonify({"error": str(exc), "code": "BATCH_LOCKED"}), 409
    except NotFoundError as exc:
        return jsonify({"error": str(exc), "code": "NOT_FOUND"}), 404
    return jsonify({"data": {"id": str(b.id), "state": b.state.value}}), 200


@opening_balance_bp.post("/api/v1/opening-batches/<bid>/rollover")
@login_required  # type: ignore[untyped-decorator]
def rollover(bid: str) -> tuple[Any, int]:
    _require_chief()
    body = request.get_json(silent=True) or {}
    try:
        b = _svc().rollover(
            UUID(bid),
            new_fiscal_year_id=UUID(str(body["fiscal_year_id"])),
            actor=UUID(str(current_user.id)),
            reason=body.get("reason") or "year roll",
        )
    except (KeyError, ValueError) as exc:
        return jsonify({"error": str(exc), "code": "INVALID_ROLLOVER"}), 422
    except BatchLockedError as exc:
        return jsonify({"error": str(exc), "code": "BATCH_LOCKED"}), 409
    except NotFoundError as exc:
        return jsonify({"error": str(exc), "code": "NOT_FOUND"}), 404
    return jsonify({"data": {"id": str(b.id), "state": b.state.value}}), 201


@opening_balance_bp.get("/api/v1/opening-batches/templates/<kind>")
@login_required  # type: ignore[untyped-decorator]
def download_template(kind: str) -> Any:
    from io import BytesIO
    from flask import send_file
    from openpyxl import Workbook

    kind = kind.lower()
    wb = Workbook()
    ws = wb.active
    if kind == "gl":
        ws.append(["account_code", "debit", "credit", "currency_code"])
        ws.append(["1111", 0, 500, "VND"])
    elif kind == "bank":
        ws.append(["bank_account_id", "amount"])
        ws.append(["", 0])
    elif kind == "counterparty":
        ws.append(["account_code", "party_id", "side", "amount", "proof"])
        ws.append(["1311", "", "debit", 0, False])
    elif kind == "stock":
        ws.append(["product_id", "warehouse_id", "qty", "total_value", "lot_code", "expiry_date", "receipt_date", "receipt_doc", "unit_cost"])
        ws.append(["", "", 0, 0, "", "", "", "", ""])
    elif kind == "assets":
        ws.append(["kind", "code", "name", "original_cost", "remaining_value", "months_left", "expense_account"])
        ws.append(["fixed_asset", "", "", 0, 0, 12, ""])
    else:
        return jsonify({"error": "unknown template", "code": "INVALID_KIND"}), 422

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(bio, as_attachment=True, download_name=f"opening_{kind}_template.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@opening_balance_bp.post("/api/v1/opening-batches/<bid>/reopen")
@login_required  # type: ignore[untyped-decorator]
def reopen(bid: str) -> tuple[Any, int]:
    role = getattr(current_user, "role", "")
    body = request.get_json(silent=True) or {}
    try:
        b = _svc().reopen(
            UUID(bid),
            actor=UUID(str(current_user.id)),
            reason=body.get("reason") or "reopen",
            is_chief=(role in ("CHIEF_ACCOUNTANT", "ADMIN")),
        )
    except PermissionError as exc:
        return jsonify({"error": str(exc), "code": "SOD_VIOLATION"}), 403
    except NotFoundError as exc:
        return jsonify({"error": str(exc), "code": "NOT_FOUND"}), 404
    return jsonify({"data": {"id": str(b.id), "state": b.state.value}}), 200
